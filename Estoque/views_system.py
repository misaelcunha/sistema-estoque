import os
import io

from flask import render_template, request, redirect, session, flash, url_for, send_file
from flask_paginate import Pagination, get_page_args
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.sql import func

from helpers import FormularioMaterial
from models import Materiais, Movimentacoes
from sistema import app, db
from upload import arquivos_permitidos
from exports import verificar_e_recriar_modelo, estilo_estoque, ajustar_largura_colunas, estilo_historico
from utils import get_materiais, get_historico, get_hora, login_required

@app.route('/')
@login_required
def index():
    # Calcula o total de materiais, materiais zerados, total de entradas e saídas
    total = Materiais.query.count()
    materiais_zerados = db.session.query(func.count(Materiais.codigo)).filter(Materiais.quantidade == 0).scalar()
    total_entradas = db.session.query(func.sum(Movimentacoes.quantidade)).filter(Movimentacoes.tipo == 'ENTRADA').scalar()
    total_saidas = db.session.query(func.sum(Movimentacoes.quantidade)).filter(Movimentacoes.tipo == 'SAÍDA').scalar()

    # Obtém as movimentações recentes
    movimentacoes_recentes = Movimentacoes.query.order_by(Movimentacoes.data_hora.desc()).limit(5).all()

    # Renderiza a página inicial com as informações calculadas
    return render_template('index.html', total=total, materiais_zerados=materiais_zerados, total_entradas=total_entradas,
                           total_saidas=total_saidas, movimentacoes=movimentacoes_recentes)

@app.route('/estoque', methods=['GET', 'POST'])
@login_required
def mostra_estoque():
    # Obtém o termo de busca e configura a paginação
    search = request.args.get('search', '').upper()
    page, per_page, offset = get_page_args(page_parameter='page', per_page_parameter='per_page')

    # Se há uma busca, filtra os materiais pelo termo de busca
    if search:
        lista_materiais = Materiais.query.filter(
            (Materiais.codigo.contains(search)) |
            (Materiais.descricao.contains(search)) |
            (Materiais.modelo.contains(search)) |
            (Materiais.fabricante.contains(search))
        ).order_by(Materiais.codigo).offset(offset).limit(per_page).all()
        total = Materiais.query.filter(
            (Materiais.codigo.contains(search)) |
            (Materiais.descricao.contains(search)) |
            (Materiais.modelo.contains(search)) |
            (Materiais.fabricante.contains(search))
        ).count()
    else:
        # Caso contrário, obtém a lista paginada de materiais
        lista_materiais = get_materiais(offset=offset, per_page=per_page)
        total = Materiais.query.count()

    # Configura a paginação
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap5')
    caminho = request.path
    return render_template('estoque.html', titulo='Materiais em Estoque', materiais=lista_materiais,
                           page=page, per_page=per_page, pagination=pagination, caminho_atual=caminho)

@app.route('/cadastro')
@login_required
def novo_material():
    # Cria uma nova instância do formulário de material
    form = FormularioMaterial()
    return render_template('cadastro_material.html', titulo='Adicionar Material', form=form)

@app.route('/criar_material', methods=['POST'])
def criar_material():
    form = FormularioMaterial(request.form)
    if form.validate_on_submit():
        # Verifica se o material já existe
        material_existente = Materiais.query.filter_by(modelo=form.modelo.data).first()

        if material_existente:
            flash('Material já cadastrado!')
            return redirect(url_for('novo_material'))

        foto = request.files.get('foto')

        # Se uma foto foi enviada e é permitida, salva a foto
        if foto and arquivos_permitidos(foto.filename):
            filename = secure_filename(foto.filename)
            foto_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            foto.save(foto_path)
            foto_filename = os.path.basename(foto_path)
        else:
            foto_filename = None

        # Cria uma nova instância de material e adiciona ao banco de dados
        novo_material = Materiais(unidade=form.unidade.data,
                                  quantidade=form.quantidade.data,
                                  descricao=form.descricao.data,
                                  modelo=form.modelo.data,
                                  fabricante=form.fabricante.data,
                                  data_hora=get_hora(),
                                  foto=foto_filename)
        db.session.add(novo_material)
        db.session.commit()

        # Registra a movimentação de adição
        movimentacao = Movimentacoes(
            material_codigo=novo_material.codigo,
            quantidade=novo_material.quantidade,
            tipo='ADICIONADO',
            descricao=novo_material.descricao,
            data_hora=novo_material.data_hora,
            feito_por=session.get('usuario_logado')
        )
        db.session.add(movimentacao)
        db.session.commit()

        flash('Material adicionado com sucesso!')
        return redirect(url_for('mostra_historico'))
    else:
        # Em caso de erro de validação, exibe os erros
        for fieldName, errorMessages in form.errors.items():
            for err in errorMessages:
                print(f'Erro no campo {fieldName}: `{err}')
        flash('Erro na validação do formulário. Por favor, verifique os campos e tente novamente.')
        return redirect(url_for('novo_material'))

@app.route('/editar_material/<int:codigo>')
@login_required
def editar_material(codigo):
    # Obtém o material pelo código e preenche o formulário com seus dados
    material = Materiais.query.filter_by(codigo=codigo).first()
    form = FormularioMaterial()
    form.unidade.data = material.unidade
    form.quantidade.data = material.quantidade
    form.descricao.data = material.descricao
    form.modelo.data = material.modelo
    form.fabricante.data = material.fabricante
    form.foto.data = material.foto

    return render_template('editar_material.html', titulo='Editar Material', material=material, form=form)

@app.route('/atualizar_material/', methods=['POST'])
def atualizar_material():
    form = FormularioMaterial(request.form)
    codigo = request.form.get('codigo')
    material = Materiais.query.filter_by(codigo=codigo).first()

    if form.validate_on_submit():
        # Atualiza os dados do material
        material.unidade = form.unidade.data
        material.quantidade = form.quantidade.data
        material.descricao = form.descricao.data
        material.modelo = form.modelo.data
        material.fabricante = form.fabricante.data
        material.data_hora = get_hora()

        # Se uma nova foto foi enviada, atualiza a foto
        nova_foto = request.files.get('foto')
        if nova_foto and arquivos_permitidos(nova_foto.filename):
            if material.foto:
                old_foto_path = os.path.join(app.config['UPLOAD_FOLDER'], material.foto)
                if os.path.exists(old_foto_path):
                    os.remove(old_foto_path)

            filename = secure_filename(nova_foto.filename)
            foto_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            nova_foto.save(foto_path)
            material.foto = filename

        db.session.add(material)
        db.session.commit()

        flash('Dados do material atualizados com sucesso!')

    else:
        # Em caso de erro de validação, exibe os erros
        for fieldName, errorMessages in form.errors.items():
            for err in errorMessages:
                print(f'Erro no campo {fieldName}: `{err}')
        flash('Erro na validação do formulário. Por favor, verifique os campos e tente novamente.')
        return redirect(url_for('editar_material'))

    return redirect(url_for('mostra_estoque'))

@app.route('/deletar_material/<string:codigo>')
@login_required
def deletar_material(codigo):
    # Obtém o usuário logado e o material a ser deletado
    usuario = session.get('usuario_logado')
    material = Materiais.query.filter_by(codigo=codigo).first()

    if material:
        # Registra a movimentação de deleção
        movimentacao = Movimentacoes(
            material_codigo=material.codigo,
            quantidade=material.quantidade,
            tipo='DELETADO',
            descricao=material.descricao,
            data_hora=get_hora(),
            feito_por=usuario
        )
        db.session.add(movimentacao)
        db.session.commit()

        # Deleta a foto associada ao material, se existir
        if material.foto and os.path.exists(material.foto):
            os.remove(material.foto)

        # Deleta o material do banco de dados
        db.session.delete(material)
        db.session.commit()
        flash('Material deletado com sucesso!')
    return redirect(url_for('mostra_historico'))

@app.route('/entrada_saida')
@login_required
def entrada_saida():
    # Cria uma nova instância do formulário de material e configura a paginação
    form = FormularioMaterial(request.form)
    search = request.args.get('search', '').upper()
    page, per_page, offset = get_page_args(page_parameter='page', per_page_parameter='per_page')

    # Se há uma busca, filtra os materiais pelo termo de busca
    if search:
        lista_materiais = Materiais.query.filter(
            (Materiais.codigo.contains(search)) |
            (Materiais.descricao.contains(search)) |
            (Materiais.modelo.contains(search)) |
            (Materiais.fabricante.contains(search))
        ).order_by(Materiais.codigo).offset(offset).limit(per_page).all()
        total = Materiais.query.filter(
            (Materiais.codigo.contains(search)) |
            (Materiais.descricao.contains(search)) |
            (Materiais.modelo.contains(search)) |
            (Materiais.fabricante.contains(search))
        ).count()
    else:
        # Caso contrário, obtém a lista paginada de materiais
        lista_materiais = get_materiais(offset=offset, per_page=per_page)
        total = Materiais.query.count()

    # Configura a paginação
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap5')
    return render_template('entrada_saida.html', titulo='Lista de Materiais', materiais=lista_materiais,
                           form=form, page=page, per_page=per_page, pagination=pagination,
                           caminho_atual=request.path)

@app.route('/entrada/<int:codigo>', methods=['POST'])
def entrada_material(codigo):
    # Obtém o material pelo código e o usuário logado
    material = Materiais.query.get(codigo)
    usuario = session.get('usuario_logado')

    if material:
        quantidade = int(request.form['quantidade'])
        if quantidade <= 0:
            flash('Por favor, digite um número positivo!')
            return redirect(url_for('entrada_saida'))
        else:
            # Atualiza a quantidade do material e registra a movimentação de entrada
            material.quantidade += quantidade
            movimentacao = Movimentacoes(
                material_codigo=codigo,
                quantidade=quantidade,
                tipo='ENTRADA',
                descricao=material.descricao,
                data_hora=get_hora(),
                feito_por=usuario
            )
            db.session.add(movimentacao)
            db.session.commit()
            flash('Entrada registrada com sucesso!')
    else:
        flash('Material não encontrado')
    return redirect(url_for('mostra_historico'))

@app.route('/saida/<int:codigo>', methods=['POST'])
def saida_material(codigo):
    # Obtém o material pelo código e o usuário logado
    material = Materiais.query.get(codigo)
    usuario = session.get('usuario_logado')
    if material:
        quantidade = int(request.form['quantidade'])
        if quantidade <= 0:
            flash('Por favor, digite um número positivo!')
            return redirect(url_for('entrada_saida'))
        else:
            if material.quantidade >= quantidade:
                # Atualiza a quantidade do material e registra a movimentação de saída
                material.quantidade -= quantidade
                movimentacao = Movimentacoes(
                    material_codigo=codigo,
                    quantidade=quantidade,
                    tipo='SAÍDA',
                    descricao=material.descricao,
                    data_hora=get_hora(),
                    feito_por=usuario
                )
                db.session.add(movimentacao)
                db.session.commit()
                flash('Saída registrada com sucesso')
            else:
                flash('Quantidade insuficiente em estoque!')
                return redirect(url_for('entrada_saida'))
    else:
        flash('Material não encontrado')
    return redirect(url_for('mostra_historico'))

@app.route('/historico')
@login_required
def mostra_historico():
    # Obtém o termo de busca e configura a paginação
    search = request.args.get('search', '').upper()
    page, per_page, offset = get_page_args(page_parameter='page', per_page_parameter='per_page')
    if search:
        # Filtra as movimentações pelo termo de busca
        movimentacoes = Movimentacoes.query.filter(
            (Movimentacoes.material_codigo.contains(search)) |
            (Movimentacoes.descricao.contains(search)) |
            (Movimentacoes.tipo.contains(search))
        ).order_by(Movimentacoes.data_hora.desc()).offset(offset).limit(per_page).all()
        total = Movimentacoes.query.filter(
            (Movimentacoes.material_codigo.contains(search)) |
            (Movimentacoes.descricao.contains(search)) |
            (Movimentacoes.tipo.contains(search))
        ).count()
    else:
        # Obtém a lista paginada de movimentações
        movimentacoes = get_historico(offset=offset, per_page=per_page)
        total = Movimentacoes.query.count()

    # Configura a paginação
    pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap5')
    return render_template('historico_movimentacao.html', titulo='Histórico de Movimentações',
                           movimentacoes=movimentacoes, page=page, per_page=per_page, pagination=pagination,
                           caminho_atual=request.path)

@app.route('/exportar_estoque')
def exportar_estoque():
    caminho_modelo = 'files/modelo.xlsx'
    caminho_modelo_original = 'files/modelo_original.xlsx'

    # Verifica e recria o arquivo modelo se necessário
    verificar_e_recriar_modelo(caminho_modelo, caminho_modelo_original)

    # Carrega o arquivo modelo
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Obtém todos os materiais
    lista_materiais = Materiais.query.all()

    # Adiciona os dados dos materiais a partir da linha 10
    linha_inicial = 10
    for i, material in enumerate(lista_materiais, start=linha_inicial):
        cells = [
            ws.cell(row=i, column=2, value=material.codigo),
            ws.cell(row=i, column=3, value=material.unidade),
            ws.cell(row=i, column=4, value=material.quantidade),
            ws.cell(row=i, column=5, value=material.descricao),
            ws.cell(row=i, column=6, value=material.modelo),
            ws.cell(row=i, column=7, value=material.fabricante),
            ws.cell(row=i, column=8, value=material.data_hora)
        ]
        for cell in cells:
            col_letter = get_column_letter(cell.column)
            estilo_estoque(cell, col_letter)

    # Ajusta a largura das colunas automaticamente
    ajustar_largura_colunas(ws)

    # Prepara o arquivo para download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='Estoque ESA.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar_movimentacoes')
def exportar_movimentacoes():
    caminho_modelo = 'files/modelo_historico.xlsx'
    caminho_modelo_original = 'files/modelo_original.xlsx'

    # Verifica e recria o arquivo modelo se necessário
    verificar_e_recriar_modelo(caminho_modelo, caminho_modelo_original)

    # Carrega o arquivo modelo
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Obtém todas as movimentações
    lista_movimentacoes = Movimentacoes.query.all()

    # Adiciona os dados dos materiais a partir da linha 10
    linha_inicial = 10
    for i, movimentacao in enumerate(lista_movimentacoes, start=linha_inicial):
        cells = [
            ws.cell(row=i, column=2, value=movimentacao.id),
            ws.cell(row=i, column=3, value=movimentacao.material_codigo),
            ws.cell(row=i, column=4, value=movimentacao.quantidade),
            ws.cell(row=i, column=5, value=movimentacao.tipo),
            ws.cell(row=i, column=6, value=movimentacao.descricao),
            ws.cell(row=i, column=7, value=movimentacao.data_hora),
            ws.cell(row=i, column=8, value=movimentacao.feito_por)
        ]
        for cell in cells:
            col_letter = get_column_letter(cell.column)
            estilo_historico(cell, col_letter)

    # Ajusta a largura das colunas automaticamente
    ajustar_largura_colunas(ws)

    # Prepara o arquivo para download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='Histórico Estoque ESA.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')